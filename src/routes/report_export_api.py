"""
報表導出系統API路由
支援PDF和Excel格式的導出
"""

from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file
from src.models.food_item import db, FoodItem
from src.models.inventory_record import InventoryRecord
from src.models.operation_log import OperationLog
import csv
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 創建藍圖
report_export_bp = Blueprint('report_export', __name__)


@report_export_bp.route('/api/reports/inventory-summary', methods=['GET'])
def export_inventory_summary():
    """
    導出庫存摘要報表
    支援PDF和Excel格式
    """
    try:
        format_type = request.args.get('format', 'pdf').lower()  # pdf or excel
        
        # 獲取庫存數據
        food_items = FoodItem.query.all()
        
        # 計算統計信息
        total_items = len(food_items)
        total_quantity = sum(item.quantity for item in food_items)
        total_value = sum(item.quantity * (item.unit_price or 0) for item in food_items)
        
        # 按分類統計
        category_stats = {}
        for item in food_items:
            cat = item.category or '未分類'
            if cat not in category_stats:
                category_stats[cat] = {'count': 0, 'quantity': 0, 'value': 0}
            category_stats[cat]['count'] += 1
            category_stats[cat]['quantity'] += item.quantity
            category_stats[cat]['value'] += item.quantity * (item.unit_price or 0)
        
        if format_type == 'excel':
            return export_inventory_excel(food_items, category_stats, total_items, total_quantity, total_value)
        else:
            return export_inventory_pdf(food_items, category_stats, total_items, total_quantity, total_value)
    except Exception as e:
        return jsonify({'error': f'導出失敗: {str(e)}'}), 500


def export_inventory_excel(food_items, category_stats, total_items, total_quantity, total_value):
    """導出庫存摘要為Excel"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '庫存摘要'
        
        # 設定列寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # 標題
        title_cell = ws['A1']
        title_cell.value = '庫存摘要報表'
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # 生成日期
        ws['A2'] = f'生成日期: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # 統計摘要
        ws['A4'] = '統計摘要'
        ws['A4'].font = Font(bold=True, size=12)
        
        ws['A5'] = '總商品數'
        ws['B5'] = total_items
        ws['A6'] = '總庫存量'
        ws['B6'] = total_quantity
        ws['A7'] = '庫存總值'
        ws['B7'] = f'${total_value:.2f}'
        
        # 分類統計表
        ws['A9'] = '按分類統計'
        ws['A9'].font = Font(bold=True, size=12)
        
        # 表頭
        headers = ['分類', '商品數', '庫存量', '庫存值', '平均單價']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        
        # 數據行
        row = 11
        for category, stats in sorted(category_stats.items()):
            ws.cell(row=row, column=1).value = category
            ws.cell(row=row, column=2).value = stats['count']
            ws.cell(row=row, column=3).value = stats['quantity']
            ws.cell(row=row, column=4).value = f'${stats["value"]:.2f}'
            avg_price = stats['value'] / stats['quantity'] if stats['quantity'] > 0 else 0
            ws.cell(row=row, column=5).value = f'${avg_price:.2f}'
            row += 1
        
        # 商品詳情表
        ws['A' + str(row + 2)] = '商品詳情'
        ws['A' + str(row + 2)].font = Font(bold=True, size=12)
        
        row += 3
        headers = ['商品名稱', '分類', '庫存量', '單位價格', '庫存值', '到期日期']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        
        row += 1
        for item in food_items:
            ws.cell(row=row, column=1).value = item.name
            ws.cell(row=row, column=2).value = item.category
            ws.cell(row=row, column=3).value = item.quantity
            ws.cell(row=row, column=4).value = f'${item.unit_price:.2f}' if item.unit_price else '-'
            value = item.quantity * (item.unit_price or 0)
            ws.cell(row=row, column=5).value = f'${value:.2f}'
            ws.cell(row=row, column=6).value = item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else '-'
            row += 1
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventory_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': f'Excel導出失敗: {str(e)}'}), 500


def export_inventory_pdf(food_items, category_stats, total_items, total_quantity, total_value):
    """導出庫存摘要為PDF"""
    try:
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # 標題
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph('庫存摘要報表', title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # 生成日期
        elements.append(Paragraph(f'生成日期: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # 統計摘要
        elements.append(Paragraph('統計摘要', styles['Heading2']))
        summary_data = [
            ['總商品數', str(total_items)],
            ['總庫存量', str(total_quantity)],
            ['庫存總值', f'${total_value:.2f}']
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # 分類統計表
        elements.append(Paragraph('按分類統計', styles['Heading2']))
        category_data = [['分類', '商品數', '庫存量', '庫存值']]
        for category, stats in sorted(category_stats.items()):
            category_data.append([
                category,
                str(stats['count']),
                str(stats['quantity']),
                f'${stats["value"]:.2f}'
            ])
        
        category_table = Table(category_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(category_table)
        elements.append(PageBreak())
        
        # 商品詳情表
        elements.append(Paragraph('商品詳情', styles['Heading2']))
        product_data = [['商品名稱', '分類', '庫存量', '單位價格', '到期日期']]
        for item in food_items:
            product_data.append([
                item.name,
                item.category or '-',
                str(item.quantity),
                f'${item.unit_price:.2f}' if item.unit_price else '-',
                item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else '-'
            ])
        
        product_table = Table(product_data, colWidths=[1.5*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(product_table)
        
        # 生成PDF
        doc.build(elements)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'inventory_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
    except Exception as e:
        return jsonify({'error': f'PDF導出失敗: {str(e)}'}), 500


@report_export_bp.route('/api/reports/sales-report', methods=['GET'])
def export_sales_report():
    """
    導出銷售報表
    """
    try:
        format_type = request.args.get('format', 'pdf').lower()
        start_date = request.args.get('start_date', None)
        end_date = request.args.get('end_date', None)
        
        # 構建查詢
        query = InventoryRecord.query.filter_by(operation_type='銷售')
        
        if start_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(InventoryRecord.created_at >= start)
            except:
                pass
        
        if end_date:
            try:
                end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(InventoryRecord.created_at < end)
            except:
                pass
        
        records = query.order_by(InventoryRecord.created_at.desc()).all()
        
        # 計算統計信息
        total_sales = len(records)
        total_amount = sum(record.total_amount or 0 for record in records)
        total_discount = sum(record.discount or 0 for record in records)
        net_amount = total_amount - total_discount
        
        if format_type == 'excel':
            return export_sales_excel(records, total_sales, total_amount, total_discount, net_amount)
        else:
            return export_sales_pdf(records, total_sales, total_amount, total_discount, net_amount)
    except Exception as e:
        return jsonify({'error': f'導出失敗: {str(e)}'}), 500


def export_sales_excel(records, total_sales, total_amount, total_discount, net_amount):
    """導出銷售報表為Excel"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '銷售報表'
        
        # 設定列寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # 標題
        title_cell = ws['A1']
        title_cell.value = '銷售報表'
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # 生成日期
        ws['A2'] = f'生成日期: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # 統計摘要
        ws['A4'] = '統計摘要'
        ws['A4'].font = Font(bold=True, size=12)
        
        ws['A5'] = '銷售筆數'
        ws['B5'] = total_sales
        ws['A6'] = '銷售總額'
        ws['B6'] = f'${total_amount:.2f}'
        ws['A7'] = '折扣總額'
        ws['B7'] = f'${total_discount:.2f}'
        ws['A8'] = '淨銷售額'
        ws['B8'] = f'${net_amount:.2f}'
        
        # 銷售詳情表
        ws['A10'] = '銷售詳情'
        ws['A10'].font = Font(bold=True, size=12)
        
        # 表頭
        headers = ['日期', '商品名稱', '分類', '數量', '銷售額', '折扣']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        
        # 數據行
        row = 12
        for record in records:
            ws.cell(row=row, column=1).value = record.created_at.strftime('%Y-%m-%d %H:%M')
            ws.cell(row=row, column=2).value = record.product_name
            ws.cell(row=row, column=3).value = record.category
            ws.cell(row=row, column=4).value = record.quantity
            ws.cell(row=row, column=5).value = f'${record.total_amount:.2f}' if record.total_amount else '-'
            ws.cell(row=row, column=6).value = f'${record.discount:.2f}' if record.discount else '-'
            row += 1
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': f'Excel導出失敗: {str(e)}'}), 500


def export_sales_pdf(records, total_sales, total_amount, total_discount, net_amount):
    """導出銷售報表為PDF"""
    try:
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # 標題
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph('銷售報表', title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # 生成日期
        elements.append(Paragraph(f'生成日期: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # 統計摘要
        elements.append(Paragraph('統計摘要', styles['Heading2']))
        summary_data = [
            ['銷售筆數', str(total_sales)],
            ['銷售總額', f'${total_amount:.2f}'],
            ['折扣總額', f'${total_discount:.2f}'],
            ['淨銷售額', f'${net_amount:.2f}']
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # 銷售詳情表
        elements.append(Paragraph('銷售詳情', styles['Heading2']))
        sales_data = [['日期', '商品名稱', '分類', '數量', '銷售額', '折扣']]
        for record in records:
            sales_data.append([
                record.created_at.strftime('%Y-%m-%d %H:%M'),
                record.product_name,
                record.category or '-',
                str(record.quantity),
                f'${record.total_amount:.2f}' if record.total_amount else '-',
                f'${record.discount:.2f}' if record.discount else '-'
            ])
        
        sales_table = Table(sales_data, colWidths=[1.2*inch, 1.5*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch])
        sales_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(sales_table)
        
        # 生成PDF
        doc.build(elements)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
    except Exception as e:
        return jsonify({'error': f'PDF導出失敗: {str(e)}'}), 500


@report_export_bp.route('/api/reports/operation-log', methods=['GET'])
def export_operation_log():
    """
    導出操作日誌報表
    """
    try:
        format_type = request.args.get('format', 'csv').lower()
        start_date = request.args.get('start_date', None)
        end_date = request.args.get('end_date', None)
        
        # 構建查詢
        query = OperationLog.query
        
        if start_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(OperationLog.created_at >= start)
            except:
                pass
        
        if end_date:
            try:
                end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(OperationLog.created_at < end)
            except:
                pass
        
        logs = query.order_by(OperationLog.created_at.desc()).all()
        
        if format_type == 'csv':
            # 創建CSV
            output = StringIO()
            writer = csv.writer(output)
            
            # 寫入頭部
            writer.writerow([
                '時間', '操作類型', '模塊', '操作描述', '對象類型', '對象名稱',
                '用戶ID', '用戶名稱', '用戶IP', '狀態', '錯誤信息'
            ])
            
            # 寫入數據
            for log in logs:
                writer.writerow([
                    log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    log.operation_type,
                    log.module or '-',
                    log.action_description or '-',
                    log.object_type or '-',
                    log.object_name or '-',
                    log.user_id or '-',
                    log.user_name or '-',
                    log.user_ip or '-',
                    log.status,
                    log.error_message or '-'
                ])
            
            output.seek(0)
            return send_file(
                BytesIO(output.getvalue().encode('utf-8-sig')),
                mimetype='text/csv; charset=utf-8',
                as_attachment=True,
                download_name=f'operation_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )
        else:
            # Excel格式
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '操作日誌'
            
            # 表頭
            headers = ['時間', '操作類型', '模塊', '操作描述', '對象類型', '對象名稱', '用戶ID', '用戶名稱', '用戶IP', '狀態']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
            
            # 數據行
            row = 2
            for log in logs:
                ws.cell(row=row, column=1).value = log.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row, column=2).value = log.operation_type
                ws.cell(row=row, column=3).value = log.module or '-'
                ws.cell(row=row, column=4).value = log.action_description or '-'
                ws.cell(row=row, column=5).value = log.object_type or '-'
                ws.cell(row=row, column=6).value = log.object_name or '-'
                ws.cell(row=row, column=7).value = log.user_id or '-'
                ws.cell(row=row, column=8).value = log.user_name or '-'
                ws.cell(row=row, column=9).value = log.user_ip or '-'
                ws.cell(row=row, column=10).value = log.status
                row += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'operation_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
    except Exception as e:
        return jsonify({'error': f'導出失敗: {str(e)}'}), 500
